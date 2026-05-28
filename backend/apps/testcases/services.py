"""
测试用例导入导出服务
支持 Excel 和 JSON 格式
"""
import json
from typing import Any

from openpyxl import Workbook, load_workbook

from apps.testcases.models import TestCase


class TestCaseImportExportService:
    """测试用例导入导出服务"""
    
    # Excel 表头
    EXCEL_HEADERS = [
        "用例标题",
        "用例描述",
        "前置条件",
        "测试步骤(JSON)",
        "预期结果",
        "优先级",
        "状态",
    ]
    
    # Excel 列映射
    EXCEL_COLUMN_MAPPING = {
        "用例标题": "title",
        "用例描述": "description",
        "前置条件": "precondition",
        "测试步骤(JSON)": "steps",
        "预期结果": "expected_result",
        "优先级": "priority",
        "状态": "status",
    }
    
    @classmethod
    def export_to_excel(cls, testcases: list[TestCase], file_path: str) -> dict[str, Any]:
        """
        导出测试用例到 Excel 文件
        
        Args:
            testcases: 测试用例查询集或列表
            file_path: 文件保存路径
            
        Returns:
            包含导出结果的字典
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"
        
        # 写入表头
        ws.append(cls.EXCEL_HEADERS)
        
        # 写入数据
        for case in testcases:
            row = [
                case.title,
                case.description,
                case.precondition,
                json.dumps(case.steps, ensure_ascii=False),
                case.expected_result,
                case.priority,
                case.status,
            ]
            ws.append(row)
        
        wb.save(file_path)
        wb.close()
        
        return {
            "success": True,
            "count": len(testcases),
            "file_path": file_path,
        }
    
    @classmethod
    def export_to_json(cls, testcases: list[TestCase]) -> dict[str, Any]:
        """
        导出测试用例到 JSON 格式
        
        Args:
            testcases: 测试用例查询集或列表
            
        Returns:
            包含导出结果的字典和 JSON 数据
        """
        data = []
        for case in testcases:
            data.append({
                "title": case.title,
                "description": case.description,
                "precondition": case.precondition,
                "steps": case.steps,
                "expected_result": case.expected_result,
                "priority": case.priority,
                "status": case.status,
            })
        
        return {
            "success": True,
            "count": len(data),
            "data": data,
        }
    
    @classmethod
    def import_from_excel(cls, file_path: str, project_id: int, user) -> dict[str, Any]:
        """
        从 Excel 文件导入测试用例
        
        Args:
            file_path: Excel 文件路径
            project_id: 所属项目ID
            user: 当前用户
            
        Returns:
            包含导入结果的字典
        """
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 验证表头
        headers = [cell.value for cell in ws[1]]
        if headers[: len(cls.EXCEL_HEADERS)] != cls.EXCEL_HEADERS:
            return {
                "success": False,
                "error": "Excel 格式不正确，请使用标准模板",
            }
        
        imported_count = 0
        error_rows = []
        
        # 从第二行开始读取数据
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # 跳过空行
                continue
            
            try:
                # 解析数据
                title = row[0]
                description = row[1] or ""
                precondition = row[2] or ""
                
                # 解析步骤 JSON
                steps_str = row[3]
                if steps_str:
                    try:
                        steps = json.loads(steps_str)
                    except json.JSONDecodeError:
                        steps = []
                else:
                    steps = []
                
                expected_result = row[4] or ""
                priority = row[5] if row[5] in ["P0", "P1", "P2", "P3"] else "P2"
                status = row[6] if row[6] in ["draft", "active", "archived"] else "draft"
                
                # 创建测试用例
                TestCase.objects.create(
                    project_id=project_id,
                    title=title,
                    description=description,
                    precondition=precondition,
                    steps=steps,
                    expected_result=expected_result,
                    priority=priority,
                    status=status,
                    created_by=user,
                )
                imported_count += 1
                
            except Exception as e:
                error_rows.append({"row": row_idx, "error": str(e)})
        
        wb.close()
        
        return {
            "success": True,
            "imported_count": imported_count,
            "error_rows": error_rows,
        }
    
    @classmethod
    def import_from_json(cls, json_data: list, project_id: int, user) -> dict[str, Any]:
        """
        从 JSON 数据导入测试用例
        
        Args:
            json_data: JSON 数据列表
            project_id: 所属项目ID
            user: 当前用户
            
        Returns:
            包含导入结果的字典
        """
        if not isinstance(json_data, list):
            return {
                "success": False,
                "error": "JSON 数据必须是数组格式",
            }
        
        imported_count = 0
        error_items = []
        
        for idx, item in enumerate(json_data):
            try:
                if not item.get("title"):
                    error_items.append({"index": idx, "error": "缺少用例标题"})
                    continue
                
                TestCase.objects.create(
                    project_id=project_id,
                    title=item["title"],
                    description=item.get("description", ""),
                    precondition=item.get("precondition", ""),
                    steps=item.get("steps", []),
                    expected_result=item.get("expected_result", ""),
                    priority=item.get("priority", "P2"),
                    status=item.get("status", "draft"),
                    created_by=user,
                )
                imported_count += 1
                
            except Exception as e:
                error_items.append({"index": idx, "error": str(e)})
        
        return {
            "success": True,
            "imported_count": imported_count,
            "error_items": error_items,
        }
